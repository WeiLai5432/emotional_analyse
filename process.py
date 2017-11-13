# coding=utf-8

import jieba
import openpyxl
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import jieba.posseg as pseg

def sentence(sen):
	jieba.load_userdict("newdic.txt")
	jieba.enable_parallel(4) # 
	# pseg = jieba.posseg
	c = pseg.cut(sen)
	res = {}
	mc = ''
	for word, flag in c:
		# print word, flag
		if flag == u'n':
			mc = word

		if flag == u'a':
			if mc != '':
			# if flag == u'a':
				res[word] = mc
				# print word, mc

				# res[mc] = word
				mc = ''
			else:
				res[word] = 'NULL'
	return res


if __name__ == '__main__':

	jieba.load_userdict("newdic.txt")
	# t = '麻蛋，冰箱是坏的。我已经提交了几天售后了，不理人。千万别买他家的，又小，还是坏的！'
	# print sentence(t)
	# mp = open('newmap.txt','r')
	mp = {}
	with open('plms.txt','r') as f:
		for line in f:
			temp = line.split(' ')
			mp[temp[0]] = temp[1].split('\n')[0]
	val = open('val.txt','wb')
	# val = load_workbook('val.xlsx')

	string = u'row_id\tcontent-评论内容\ttheme-主题\tsentiment_word-情感关键词\tsentiment_anls-情感正负面'
	# val.writelines(string.encode('utf-8'))

	with open('val.txt','a') as f:
		f.write(string.encode('utf-8'))

	trainfile = load_workbook('vald.xlsx')
	ws = trainfile.get_sheet_by_name(trainfile.get_sheet_names()[0])

	rows = ws.rows
	col = ws.columns

	for i, row in enumerate(rows):
		print i
		for j, cell in enumerate(row):
			if j == 0:
				string = str(i)
			if j == 1 and cell.value != None:
				# print cell.value
				sen = unicode(cell.value)
				res = sentence(sen)
				# sheet['A' + str(i)] = sen
				A = sen
				B = ''
				# print len(res)
				for ii in res.keys():
					if len(B) != 0:
						B = B + ';' + res[ii]
					else:
						B = res[ii]
						# print B

				# sheet['B' + str(i)] = B  #json.dumps(res.keys(), ensure_ascii=False)

				C = ''
				D = ''
				for ii in res.keys():
					if len(C) != 0:
						C = C + ';' + ii
						if ii.encode('utf-8') in mp.keys():
							D = D + ';' + mp[ii.encode('utf-8')]
						else:
							D = D + ';' + u'正负未知'
					else:
						C = ii
						if ii.encode('utf-8') in mp.keys():
							D = mp[ii.encode('utf-8')]
						else:
							D = u'正负未知'

				# str(res.keys()).encode('gbk')
				# sheet['C' + str(i)] = C #str(res.values()).encode('gbk')
				# val.write(unicode((sen + '\t' + str(res.keys()).encode('utf-8') + '\t' + str(res.values()) + '\n')).encode('gbk'))

				string = A + '\t' + B + '\t' + C + '\t' + D + '\n'

				# val.writelines(string.encode('utf-8'))
				with open('val.txt','a') as f:
					f.write(string.encode('utf-8'))

		# wb.save('val.csv')


